"""Generate member-only map datasets from the business sources supplied by Label Vanlife."""

from __future__ import annotations

import html
import difflib
import json
import re
import time
import unicodedata
import urllib.parse
import urllib.request
import zipfile
from pathlib import Path

import openpyxl
from lxml import html as lxml_html
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
BIENVENUE_ZIP = Path(
    r"C:\Users\cleme\Desktop\label vanlife\FICHIERS data externes\Bienvenue a la ferme.zip"
)
OUTPUT = ROOT / "src" / "data" / "member-camping-network.json"
PAPA_PDF = Path(
    r"C:\Users\cleme\Desktop\label vanlife\FICHIERS data externes\Papa'rtenaires 07 2026.pdf"
)
GEOCODE_CACHE = ROOT / "scripts" / "papa-geocode-cache.json"
KML_DATA = ROOT / "src" / "data" / "campings_kml.json"
BIENVENUE_XLSX = Path(
    r"C:\Users\cleme\Desktop\label vanlife\contacts_bienvenue_ferme_colonnes_propres.xlsx"
)


def slugify(value: str) -> str:
    value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode()
    value = re.sub(r"[^a-zA-Z0-9]+", "-", value).strip("-").lower()
    return value


def decode_json_string(value: str) -> str:
    try:
        return json.loads(f'"{value}"')
    except json.JSONDecodeError:
        return value.replace(r"\/", "/")


def extract(pattern: str, source: str, default: str = "") -> str:
    match = re.search(pattern, source)
    if not match:
        return default
    return html.unescape(decode_json_string(match.group(1)))


def compact_text(value: str) -> str:
    return " ".join(html.unescape(value).split())


def unique_text(values: list[str]) -> list[str]:
    result: list[str] = []
    for value in values:
        cleaned = compact_text(value)
        if cleaned and cleaned not in result:
            result.append(cleaned)
    return result


def extract_bienvenue_points() -> list[dict[str, object]]:
    points: list[dict[str, object]] = []
    workbook = openpyxl.load_workbook(BIENVENUE_XLSX, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    contacts = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        _, place_name, address, _, _, website = row
        if place_name:
            contacts[slugify(str(place_name))] = {
                "address": str(address or ""),
                "website": str(website or ""),
            }

    with zipfile.ZipFile(BIENVENUE_ZIP) as archive:
        for filename in archive.namelist():
            source = archive.read(filename).decode("utf-8", "ignore")
            document = lxml_html.fromstring(source)
            coordinates = re.search(
                r'\\?"coordinates\\?":\{\\?"lat\\?":(-?\d+(?:\.\d+)?),'
                r'\\?"lon\\?":(-?\d+(?:\.\d+)?)\}',
                source,
            )
            if not coordinates:
                continue
            latitude = float(coordinates.group(1))
            longitude = float(coordinates.group(2))
            # One supplied BOTALI page contains a duplicated leading digit (442.x instead of 42.x).
            if abs(latitude) > 90 and len(str(abs(latitude))) > 2:
                latitude = float(str(latitude)[1:])

            name = filename.rsplit(" - ", 3)[0]
            filename_parts = filename.rsplit(" - ", 3)
            city = filename_parts[-3] if len(filename_parts) >= 4 else ""
            department = filename_parts[-2] if len(filename_parts) >= 4 else ""
            contact = contacts.get(slugify(name), {})
            address = str(contact.get("address", ""))
            website = str(contact.get("website", ""))
            if not address:
                embedded_address = re.search(r'\\"address\\":\\"(.*?)\\"', source)
                embedded_locality = re.search(r'\\"locality\\":\\"(.*?)\\"', source)
                embedded_postal = re.search(r'\\"postalCode\\":\\"(.*?)\\"', source)
                if embedded_address:
                    address = decode_json_string(embedded_address.group(1))
                    locality = decode_json_string(embedded_locality.group(1)) if embedded_locality else ""
                    embedded_code = decode_json_string(embedded_postal.group(1)) if embedded_postal else ""
                    address = ", ".join(part for part in (address, locality, embedded_code, city) if part)
            postal_match = re.search(r"\b(\d{5})\b", address)
            postal_code = postal_match.group(1) if postal_match else ""
            description_nodes = document.xpath('//*[contains(concat(" ", normalize-space(@class), " "), " Description-text ")]')
            description = compact_text(description_nodes[0].text_content()) if description_nodes else ""
            contact_nodes = document.xpath('//*[contains(concat(" ", normalize-space(@class), " "), " Description-contacts ")]')
            contact_name = compact_text(contact_nodes[0].text_content()) if contact_nodes else ""
            phones = unique_text([
                value.removeprefix("tel:")
                for value in document.xpath('//a[starts-with(@href, "tel:")]/@href')
            ])
            emails = unique_text([
                value.removeprefix("mailto:").split("?", 1)[0]
                for value in document.xpath('//a[starts-with(@href, "mailto:")]/@href')
            ])
            images = unique_text(document.xpath(
                '//section[@id="description-slider"]//img/@src | '
                '//*[contains(@class, "ServiceLayout-description")]/following::img/@src'
            ))[:10]
            service_descriptions = unique_text([
                node.text_content()
                for node in document.xpath('//*[contains(@class, "ServiceLayout-description")]')
            ])

            points.append(
                {
                    "id": f"bienvenue-ferme-{slugify(name)}",
                    "name": name,
                    "network": "Bienvenue à la ferme",
                    "address": address,
                    "postalCode": postal_code,
                    "city": city,
                    "region": department,
                    "lat": latitude,
                    "lng": longitude,
                    "website": website or None,
                    "description": description or None,
                    "contactName": contact_name or None,
                    "phones": phones,
                    "emails": emails,
                    "images": images,
                    "details": service_descriptions,
                    "source": "Fiches Bienvenue à la ferme fournies par Label Vanlife",
                }
            )
    return points


def papa_name(text: str) -> str:
    head = re.split(r"🌟|\s+Loisirs|\s+Hébergements|\s+Accueil|\s+Services", text, maxsplit=1)[0]
    head = re.sub(r"^\d+\.\s*", "", head).strip()
    head = re.sub(r"\s+\d+\.\s+\d.*$", "", head).strip()
    return re.sub(r"\s+z$", "", head).strip()


def papa_address(text: str) -> str:
    before_phone = text.split("", 1)[0]
    postal_codes = list(re.finditer(r"\b\d{5}\b", before_phone))
    if not postal_codes:
        return ""
    postal = postal_codes[-1]
    value = before_phone[max(0, postal.start() - 125) : postal.end() + 65]
    starts = list(re.finditer(
        r"(?:\b\d+[A-Za-z]?[, .-]+(?:route|rue|avenue|allée|chemin|impasse)\b|"
        r"\b(?:lieu[- ]dit|lieudit|château|étang|l'île|le bout|le champ|ushia|claux|plaine|quartier)\b)",
        value,
        re.IGNORECASE,
    ))
    if starts:
        value = value[starts[-1].start() :]
    value = re.sub(r"^(?:.*?)(?=\b\d+[A-Za-z]?[, .-]+(?:route|rue|avenue|allée|chemin|impasse)\b)", "", value, flags=re.IGNORECASE)
    return value.strip(" .-")


def geocode(query: str, cache: dict[str, dict[str, object]]) -> dict[str, object] | None:
    if query in cache:
        return cache[query]
    params = urllib.parse.urlencode({"q": f"{query}, France", "format": "jsonv2", "limit": 1, "countrycodes": "fr"})
    request = urllib.request.Request(
        f"https://nominatim.openstreetmap.org/search?{params}",
        headers={"User-Agent": "LabelVanlife-map-import/1.0 (contact@labelvanlife.com)"},
    )
    with urllib.request.urlopen(request, timeout=20) as response:
        results = json.load(response)
    result = {"lat": float(results[0]["lat"]), "lng": float(results[0]["lon"])} if results else None
    cache[query] = result or {}
    GEOCODE_CACHE.write_text(json.dumps(cache, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    time.sleep(1.05)
    return result


def lookup_kml(name: str, campings: list[dict[str, str]]) -> dict[str, float] | None:
    target = slugify(re.sub(r"\bcamping\b|municipal", "", name, flags=re.IGNORECASE)).replace("-", "")
    best_score = 0.0
    best: dict[str, str] | None = None
    for camping in campings:
        candidate = slugify(re.sub(r"\bcamping\b|municipal", "", camping["name"], flags=re.IGNORECASE)).replace("-", "")
        score = difflib.SequenceMatcher(None, target, candidate).ratio()
        if target in candidate or candidate in target:
            score = max(score, 0.93 if min(len(target), len(candidate)) >= 8 else score)
        if score > best_score:
            best_score, best = score, camping
    if not best or best_score < 0.9:
        return None
    lng, lat, *_ = best["coordinates"].split(",")
    return {"lat": float(lat), "lng": float(lng)}


def extract_papa_points() -> list[dict[str, object]]:
    cache = json.loads(GEOCODE_CACHE.read_text(encoding="utf-8")) if GEOCODE_CACHE.exists() else {}
    campings = json.loads(KML_DATA.read_text(encoding="utf-8"))
    points: list[dict[str, object]] = []
    reader = PdfReader(PAPA_PDF)
    ignored = ("Changement de région", "Bretagne", "Bourgogne", "Centre Val", "Grand-Est", "Hauts de", "Nouvelle Aquitaine", "Occitanie", "Pays de", "* Provence")
    for page_number, page in enumerate(reader.pages[5:], start=6):
        text = " ".join((page.extract_text() or "").split())
        if not text or text.startswith(ignored):
            continue
        name = papa_name(text)
        address = papa_address(text)
        if not name or not address:
            print(f"Skipping Papa’rtenaire page {page_number}: missing name/address")
            continue
        coordinates = geocode(f"{name}, {address}", cache) or geocode(address, cache) or lookup_kml(name, campings)
        if not coordinates:
            print(f"Skipping Papa’rtenaire page {page_number}: address not geocoded ({address})")
            continue
        postal_match = re.search(r"\b(\d{5})\b", address)
        website_match = re.search(r"https?://\S+|www\.\S+", text)
        offer_match = re.search(r"Offre\s*:?(.*?)(?:Date d’ouverture|Services|$)", text, re.IGNORECASE)
        phone_match = re.search(r"\uf028\s*(.*?)(?=https?://|www\.|Loisirs|Hébergements|Accueil|$)", text)
        activities_match = re.search(r"Loisirs\s*&\s*activités\s*(.*?)(?=Hébergements\s*&\s*emplacements|Accueil|$)", text, re.IGNORECASE)
        capacity_match = re.search(r"Hébergements\s*&\s*emplacements\s*:?\s*(.*?)(?=Accueil|Date d’ouverture|Offre|$)", text, re.IGNORECASE)
        opening_match = re.search(r"Date d’ouverture\s*:\s*(.*?)(?=Offre|$)", text, re.IGNORECASE)
        activities = unique_text(re.split(r"\s*•\s*", activities_match.group(1))) if activities_match else []
        points.append({
            "id": f"papa-rtenaires-{slugify(name)}",
            "name": name,
            "network": "Papa’rtenaires",
            "address": address,
            "postalCode": postal_match.group(1) if postal_match else "",
            "city": address.split()[-1] if address else "",
            "region": "",
            "lat": coordinates["lat"],
            "lng": coordinates["lng"],
            "website": website_match.group(0).rstrip(".,)") if website_match else None,
            "memberOffer": offer_match.group(1).strip() if offer_match else "Voir les conditions du réseau",
            "phone": compact_text(phone_match.group(1)) if phone_match else None,
            "activities": activities,
            "capacity": compact_text(capacity_match.group(1)) if capacity_match else None,
            "openingHours": compact_text(opening_match.group(1)) if opening_match else None,
            "source": f"Guide Papa’rtenaires juillet 2026, page {page_number}",
        })
    return points


def main() -> None:
    points = extract_bienvenue_points() + extract_papa_points()
    OUTPUT.write_text(
        json.dumps(points, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    print(f"Generated {len(points)} verified member-only map points in {OUTPUT}")


if __name__ == "__main__":
    main()
